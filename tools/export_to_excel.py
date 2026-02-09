"""
WAT Framework — Export Leads to Excel
Writes or appends enriched leads to an .xlsx file using openpyxl.
De-duplicates by Name + Location to avoid repeat entries on re-runs.

Usage (called by the agent):
    from tools.export_to_excel import export_leads
    export_leads(leads, output_path="output/leads.xlsx")
"""

from __future__ import annotations

import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from tools.utils import PROJECT_ROOT

# ── Column definitions (order matters) ──────────────────────────────────────
COLUMNS = [
    "Name",
    "Category",
    "Location",
    "City",
    "State",
    "Phone No.",
    "Website",
    "Email",
    "Business Info",
]

# Map column header → lead dict key
FIELD_MAP = {
    "Name": "name",
    "Category": "category",
    "Location": "location",
    "City": "city",
    "State": "state",
    "Phone No.": "phone",
    "Website": "website",
    "Email": "email",
    "Business Info": "business_info",
}


def export_leads(
    leads: list[dict],
    output_path: str = "output/leads.xlsx",
) -> Path:
    """
    Write leads to a fresh Excel file containing only the current run's data.
    Any existing file at the same path is overwritten.
    Duplicates within the current batch (by Name + Location) are still skipped.

    Args:
        leads:       List of lead dicts with keys matching FIELD_MAP values.
        output_path: Path relative to project root, or an absolute path.

    Returns:
        Absolute Path to the written Excel file.
    """
    # Validate input
    if leads is None:
        raise ValueError("leads must not be None")
    if not isinstance(leads, list):
        raise TypeError(f"leads must be a list, got {type(leads).__name__}")
    for i, item in enumerate(leads):
        if not isinstance(item, dict):
            raise TypeError(f"leads[{i}] must be a dict, got {type(item).__name__}")

    # Resolve path
    path = Path(output_path)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    path.parent.mkdir(parents=True, exist_ok=True)

    # Always create a fresh workbook — only current run data
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    _write_header(ws)
    print(f"[export] Creating fresh file: {path}")

    # Write leads (deduplicate within this batch only)
    added = 0
    seen_keys = set()
    for lead in leads:
        key = _dedup_key(lead)
        if key in seen_keys:
            continue

        row = [lead.get(FIELD_MAP[col], "") for col in COLUMNS]
        ws.append(row)
        seen_keys.add(key)
        added += 1

    # Auto-fit column widths (approximate)
    _auto_fit_columns(ws)

    # Save with retry — handles "file open in Excel" PermissionError
    _save_with_retry(wb, path)
    print(f"[export] Saved: {path}  ({added} leads written)")

    return path


# ── Internal helpers ─────────────────────────────────────────────────────────

def _write_header(ws) -> None:
    """Write styled header row."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze the header row
    ws.freeze_panes = "A2"


def _auto_fit_columns(ws) -> None:
    """Set reasonable column widths based on content."""
    # Predefined widths for each column (looks better than pure auto-fit)
    widths = {
        "Name": 30,
        "Category": 20,
        "Location": 40,
        "City": 18,
        "State": 12,
        "Phone No.": 18,
        "Website": 35,
        "Email": 30,
        "Business Info": 60,
    }
    for col_idx, col_name in enumerate(COLUMNS, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(col_name, 20)


def _save_with_retry(wb, path: Path, max_attempts: int = 3) -> None:
    """Save workbook with retries for PermissionError (file open in Excel)."""
    for attempt in range(1, max_attempts + 1):
        try:
            wb.save(path)
            return
        except PermissionError:
            if attempt == max_attempts:
                raise PermissionError(
                    f"Cannot save to {path}. "
                    f"Please close the file in Excel and try again."
                )
            wait = 3 * attempt  # exponential-ish: 3s, 6s, 9s
            print(f"[export] File is locked (open in Excel?). "
                  f"Retrying in {wait}s... ({attempt}/{max_attempts})")
            time.sleep(wait)


def _dedup_key(lead: dict) -> str:
    """Create a normalized de-duplication key from name + location."""
    name = (lead.get("name") or "").strip().lower()
    location = (lead.get("location") or "").strip().lower()
    return f"{name}||{location}"





# ── CLI entry point for testing ──────────────────────────────────────────────
if __name__ == "__main__":
    sample_leads = [
        {
            "name": "Joe's Pizza",
            "category": "Pizza restaurant",
            "location": "7 Carmine St, New York, NY 10014",
            "city": "New York",
            "state": "NY",
            "phone": "(212) 366-1182",
            "website": "https://www.joespizzanyc.com",
            "email": "info@joespizzanyc.com",
            "business_info": "Joe's Pizza is an iconic New York-style pizzeria in Greenwich Village, known for its classic thin-crust slices since 1975.",
        },
        {
            "name": "Quick Fix Plumbing",
            "category": "Plumber",
            "location": "123 Main St, Chicago, IL 60601",
            "city": "Chicago",
            "state": "IL",
            "phone": "(312) 555-0100",
            "website": "",
            "email": "",
            "business_info": "Quick Fix Plumbing is a local plumbing service in Chicago offering residential and commercial plumbing repairs.",
        },
    ]

    out = export_leads(sample_leads, output_path="output/leads.xlsx")
    print(f"\nFile written to: {out}")
